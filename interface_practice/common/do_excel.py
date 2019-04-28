# -*- coding: utf-8 -*-
# @time:2019/4/17 11:21
# Author:yh
# @file:do_excel.py
from openpyxl import load_workbook

class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.check_sql = None

class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        cases = []
        for r in range(2,sheet.max_row+1):
            row_case = Case()
            row_case.case_id = sheet.cell(row=r,column=1).value
            row_case.title = sheet.cell(row=r,column=2).value
            row_case.url = sheet.cell(row=r, column=3).value
            row_case.data = sheet.cell(row=r, column=4).value
            row_case.method = sheet.cell(row=r, column=5).value
            row_case.expected = sheet.cell (row=r, column=6).value
            row_case.check_sql = sheet.cell(row=r,column=9).value
            cases.append(row_case)
        wb.close()
        return cases

    def write_back(self,row,col,value):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,col).value=value
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    do_excel = DoExcel()



